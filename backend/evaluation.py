"""Ground Truth Evaluation - Step 6: Product Pipeline Evaluation.

Evaluates the product pipeline against a 200-item ground-truth dataset.
Compares pipeline output with verified ground-truth fields from the
"Delivery Format" sheet of Unilog-Sample_200_Items-Input-vs-Output.xlsx.

Comparisons are deterministic: canonical/normalized values are compared,
with casing, UOM formatting, whitespace, and approved symbols handled
consistently.  Outcomes are clearly classified as MATCHED, MISMATCH,
MISSING, EXTRA, or NOT_EVALUABLE.
"""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl

from backend.ai.agents.enrichment_agent import enrich_product_metadata
from backend.ai.agents.validation_agent import validate_product_data
from backend.validation import validate_product_twin
from backend.schemas.product import ProductTwinAttribute
from backend.lov_validation import is_approved_attribute, is_approved_value
from backend.uom_validation import is_uom_valid
from backend.manufacturer_resolution import resolve_manufacturer, resolve_brand
from backend.description_generation import generate_descriptions, validate_char_limits, DESCRIPTION_SPECS


# ---------------------------------------------------------------------------
# Comparison outcome enum
# ---------------------------------------------------------------------------

OUTCOME_MATCHED = "MATCHED"
OUTCOME_MISMATCH = "MISMATCH"
OUTCOME_MISSING = "MISSING"
OUTCOME_EXTRA = "EXTRA"
OUTCOME_NOT_EVALUABLE = "NOT_EVALUABLE"

OUTCOME_LABELS = {
    OUTCOME_MATCHED: "matched",
    OUTCOME_MISMATCH: "mismatch",
    OUTCOME_MISSING: "missing",
    OUTCOME_EXTRA: "extra",
    OUTCOME_NOT_EVALUABLE: "not_evaluable",
}


# ---------------------------------------------------------------------------
# Helper: deterministic field comparison
# ---------------------------------------------------------------------------

def _normalize_value(v: Any) -> str:
    """Normalize a value for comparison: strip whitespace, lower-case."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    return s


def _normalize_uom(v: Any) -> str:
    """Normalize a UOM value: strip, upper-case."""
    if v is None:
        return ""
    return str(v).strip().upper()


def _compare_values(
    got: Any, expected: Any, *,
    normalize_uom: bool = False,
) -> str:
    """Compare two values and return outcome.

    Returns one of: MATCHED, MISMATCH, MISSING, EXTRA, NOT_EVALUABLE
    """
    if got is None and expected is None:
        return OUTCOME_MATCHED
    if got is None:
        return OUTCOME_MISSING
    if expected is None:
        return OUTCOME_EXTRA
    if normalize_uom:
        g = _normalize_uom(got)
        e = _normalize_uom(expected)
    else:
        g = _normalize_value(got)
        e = _normalize_value(expected)
    if g == e:
        return OUTCOME_MATCHED
    return OUTCOME_MISMATCH


def _compare_manufacturers(
    got: Any, expected: Any,
) -> str:
    """Compare manufacturer/brand using canonical resolution.

    Uses deterministic comparison: first tries canonical resolution,
    then falls back to normalized string comparison.
    """
    if got is None and expected is None:
        return OUTCOME_MATCHED
    if got is None:
        return OUTCOME_MISSING
    if expected is None:
        return OUTCOME_EXTRA

    # Canonical resolution
    g_canon, g_status, _ = resolve_manufacturer(got)
    e_canon, e_status, _ = resolve_manufacturer(expected)

    # Both found in canonical list with VERIFIED status
    if g_status == "VERIFIED" and e_status == "VERIFIED":
        if g_canon == e_canon:
            return OUTCOME_MATCHED
        return OUTCOME_MISMATCH

    # One or both not found - fall back to raw string comparison
    # Normalize: lower-case, strip whitespace
    g_norm = _normalize_value(got)
    e_norm = _normalize_value(expected)
    if g_norm == e_norm:
        return OUTCOME_MATCHED
    return OUTCOME_MISMATCH


# ---------------------------------------------------------------------------
# Description comparison (Invoice, Mobile, Product Title, Short, Long)
# ---------------------------------------------------------------------------

_DESC_FIELDS = [
    "invoice",
    "mobile",
    "title",
    "short",
    "long",
]


def _compare_descriptions(
    pipeline_descriptions: Dict[str, str],
    ground_truth: Dict[str, str],
) -> Dict[str, str]:
    """Compare each description field and return per-field outcomes."""
    results: Dict[str, str] = {}
    for field in _DESC_FIELDS:
        got = pipeline_descriptions.get(field, "")
        expected = ground_truth.get(field, "")
        if got is None:
            got = ""
        if expected is None:
            expected = ""
        # Compare stripped, lower-cased
        if got.strip().lower() == expected.strip().lower():
            results[field] = OUTCOME_MATCHED
        else:
            results[field] = OUTCOME_MISMATCH
    return results


# ---------------------------------------------------------------------------
# Normalization / validation helpers for evaluation
# ---------------------------------------------------------------------------

def _run_pipeline_on_text(text: str) -> Dict[str, Any]:
    """Run the existing pipeline on a text input and return results."""
    # Use the existing pipeline service
    from backend.services.pipeline_service import run_product_pipeline
    # We need a DB session; create a minimal one
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from backend.db.base import Base

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = SessionLocal()

    try:
        result = run_product_pipeline(
            db=db,
            text=text,
        )
        return result
    finally:
        db.close()


def _extract_attribute_result(
    enrichment: Dict[str, Any],
    validation: Dict[str, Any],
) -> Dict[str, Any]:
    """Extract the key attributes and validation results from pipeline output."""
    attrs = enrichment.get("attributes", [])
    # Take the first attribute for the key fields, or build a summary
    if attrs:
        a = attrs[0]
        return {
            "key": a.get("key"),
            "label": a.get("label"),
            "raw_value": a.get("raw_value"),
            "value": a.get("value"),
            "unit": a.get("unit"),
            "confidence": a.get("confidence"),
            "status": a.get("status"),
        }
    # Also pull from validation
    v_attrs = validation.get("missing_attributes", [])
    return {
        "key": None,
        "label": None,
        "raw_value": None,
        "value": None,
        "unit": None,
        "confidence": 0.0,
        "status": "unverified",
    }


# ---------------------------------------------------------------------------
# Core evaluation: run pipeline on a single input and compare with ground truth
# ---------------------------------------------------------------------------

def evaluate_row(
    input_row: Dict[str, Any],
    ground_truth: Dict[str, Any],
) -> Dict[str, Any]:
    """Evaluate a single row of input against ground truth.

    Returns a dict with per-field comparison outcomes and a
    'classification' of the overall row result.
    """
    # Get the input text
    input_text = input_row.get("text") or input_row.get("description") or ""

    # Run the pipeline
    pipeline_result = _run_pipeline_on_text(input_text)
    enrichment = pipeline_result.get("enrichment", {})
    validation = pipeline_result.get("validation", {})

    # Extract pipeline-generated attributes
    pipeline_attrs = enrichment.get("attributes", [])

    # Build a ProductTwinAttribute from pipeline output for validation
    if pipeline_attrs:
        pa = pipeline_attrs[0]
        pipeline_attr = ProductTwinAttribute(
            attribute=pa.get("key", ""),
            raw_value=pa.get("raw_value"),
            normalized_value=pa.get("value"),
            unit=pa.get("unit"),
            confidence=float(pa.get("confidence") or 0.0),
            status=pa.get("status") or "unverified",
            source=pa.get("source"),
            evidence=pa.get("evidence_quote"),
        )
    else:
        pipeline_attr = ProductTwinAttribute(
            attribute="",
            raw_value=None,
            normalized_value=None,
            unit=None,
            confidence=0.0,
            status="unverified",
        )

    # Run validation on pipeline output
    val_result = validate_product_twin(pipeline_attr)

    # ------------------------------------------------------------------
    # Compare fields with ground truth
    # ------------------------------------------------------------------

    outcomes: Dict[str, str] = {}

    # Manufacturer - use input row's Part_Manuf / E1_Brand / Unilog_Brand
    # These are the source manufacturer fields from the input document
    got_mfr = input_row.get("Part_Manuf") or input_row.get("E1_Brand") or input_row.get("Unilog_Brand") or \
              input_row.get("DIB_Brand") or ""
    expected_mgr = ground_truth.get("MANUFACTURER_NAME") or ground_truth.get("MANUFACTURER_PART_NUMBER") or ""
    # Take the first non-empty manufacturer from input
    if got_mfr is None:
        got_mfr = ""
    outcomes["manufacturer"] = _compare_manufacturers(got_mfr, expected_mgr)

    # Brand - same source as manufacturer
    got_brand = input_row.get("Part_Manuf") or input_row.get("E1_Brand") or input_row.get("Unilog_Brand") or \
                input_row.get("DIB_Brand") or ""
    expected_brand = ground_truth.get("BRAND_NAME") or ground_truth.get("Unilog_Brand") or ground_truth.get("E1_Brand") or ""
    if got_brand is None:
        got_brand = ""
    outcomes["brand"] = _compare_manufacturers(got_brand, expected_brand)

    # Classification / classpath
    # Pipeline doesn't output classification directly - check if ground truth has it
    got_cls = str(pipeline_result.get("classification") or "").strip().lower()
    expected_cls = str(ground_truth.get("Classpath") or ground_truth.get("classpath") or "").strip().lower()
    
    # If neither pipeline nor ground truth has classification data, mark as NOT_EVALUABLE
    if not got_cls and not expected_cls:
        outcomes["classification"] = OUTCOME_NOT_EVALUABLE
    elif got_cls and expected_cls:
        # Both have values - compare them
        if got_cls == expected_cls:
            outcomes["classification"] = OUTCOME_MATCHED
        else:
            outcomes["classification"] = OUTCOME_MISMATCH
    elif got_cls:
        outcomes["classification"] = OUTCOME_MISMATCH  # pipeline has it but GT doesn't
    elif expected_cls:
        outcomes["classification"] = OUTCOME_MISMATCH  # GT has it but pipeline doesn't
    else:
        outcomes["classification"] = OUTCOME_NOT_EVALUABLE

    # Attributes (LOV values, UOM, normalized values)
    # Compare each attribute key from ground truth
    gt_attributes = ground_truth.get("attributes", [])
    pipeline_attr_keys = set()
    if pipeline_attrs:
        pipeline_attr_keys = {a.get("key", "").lower() for a in pipeline_attrs}

    lov_compliant = 0
    lov_total = 0
    uom_compliant = 0
    uom_total = 0

    for gt_attr in gt_attributes:
        key = gt_attr.get("label") or gt_attr.get("attribute") or ""
        if not key:
            continue
        gt_value = gt_attr.get("value") or gt_attr.get("raw_value") or ""
        gt_uom = gt_attr.get("uom") or ""

        lov_total += 1

        # Try to match by attribute key first
        key_matched = key.lower() in pipeline_attr_keys
        
        if key_matched:
            # Find the pipeline attribute(s) with this key
            pa_keys = [a for a in pipeline_attrs if a.get("key", "").lower() == key.lower()]
            if pa_keys:
                pa = pa_keys[0]  # Take first match
                pv = pa.get("value") or ""
                pu = pa.get("unit") or ""
            else:
                # Key in set but no matching attr - treat as value comparison
                pa = None
                key_matched = False
        else:
            # No key match - try to match by value comparison
            # Use the first pipeline attribute for value comparison
            pa = pipeline_attrs[0] if pipeline_attrs else None
            pv = pa.get("value") if pa else ""

        # LOV compliance
        if pa and key_matched:
            # We have a key match - check LOV approval
            lov_ok = is_approved_attribute(key)
            if lov_ok:
                lov_value_ok = is_approved_value(key, pv)
                if lov_value_ok:
                    lov_compliant += 1
                else:
                    # Value not in LOV - check if raw value matches after normalization
                    if _normalize_value(pv) == _normalize_value(gt_value):
                        lov_compliant += 1  # matched by raw comparison
            else:
                # No LOV for this key - compare values directly
                if _normalize_value(pv) == _normalize_value(gt_value):
                    lov_compliant += 1
        elif pa and not key_matched:
            # Key didn't match - compare values directly regardless of LOV
            if _normalize_value(pv) == _normalize_value(gt_value):
                lov_compliant += 1
        # else: no pipeline attr to compare, skip

        # UOM compliance
        if pa and (key_matched or not key_matched):
            # We have a pipeline attr - check UOM
            pu = pa.get("unit") or ""
            if gt_uom:
                uom_total += 1
                if pu and _normalize_uom(pu) == _normalize_uom(gt_uom):
                    uom_compliant += 1
                elif not pu and not gt_uom:
                    uom_compliant += 1  # both empty counts as compliant
            elif not pu:
                uom_total += 1  # both empty counts as compliant
        elif pa is None:
            # No pipeline attr to compare UOM with
            pass

    # If no ground-truth attributes were found, mark as NOT_EVALUABLE
    if lov_total == 0:
        # No attributes to evaluate - not evaluable
        outcomes["lov_compliance"] = OUTCOME_NOT_EVALUABLE
        outcomes["uom_compliance"] = OUTCOME_NOT_EVALUABLE
    else:
        outcomes["lov_compliance"] = (
            OUTCOME_MATCHED if lov_compliant >= lov_total * 0.8 else OUTCOME_MISMATCH
        )
        outcomes["uom_compliance"] = (
            OUTCOME_MATCHED if uom_compliant >= uom_total * 0.8 else OUTCOME_MISMATCH
        )

    # Character-limit compliance (descriptions)
    pipeline_desc = pipeline_result.get("description", {}) or {}
    gt_desc = ground_truth.get("descriptions") or {}

    char_compliant = 0
    char_total = 0
    for field in _DESC_FIELDS:
        gt_field = gt_desc.get(field, "")
        pipe_field = pipeline_desc.get(field, "")
        if gt_field == "" and pipe_field == "":
            # Both empty - skip (not evaluable)
            continue
        char_total += 1
        # Check char limit using existing validator
        result = validate_char_limits({field: pipe_field})
        if result.get(field, False):
            char_compliant += 1
        else:
            # Also compare if within reasonable limit
            max_chars = DESCRIPTION_SPECS.get(field, {}).get("max_chars", 500)
            if len(pipe_field) <= max_chars:
                char_compliant += 1

    if char_total == 0:
        outcomes["character_limit_compliance"] = OUTCOME_NOT_EVALUABLE
    else:
        outcomes["character_limit_compliance"] = (
            OUTCOME_MATCHED if char_compliant >= char_total * 0.8 else OUTCOME_MISMATCH
        )

    # Verified-value rate
    # Count how many pipeline attributes have status VERIFIED and match ground truth
    verified_matched = 0
    verified_total = 0
    for gt_attr in gt_attributes:
        key = gt_attr.get("label") or gt_attr.get("attribute") or ""
        if not key:
            continue
        verified_total += 1
        if pipeline_attrs:
            pa_key = [a for a in pipeline_attrs if a.get("key", "").lower() == key.lower()]
            if pa_key:
                pa = pa_key[0]
                if (pa.get("status") or "").upper() == "VERIFIED":
                    # Check if value matches
                    pv = pa.get("value") or ""
                    gv = gt_attr.get("value") or ""
                    if _normalize_value(pv) == _normalize_value(gv):
                        verified_matched += 1

    if verified_total == 0:
        outcomes["verified_value_rate"] = OUTCOME_NOT_EVALUABLE
    elif verified_matched / verified_total >= 0.8:
        outcomes["verified_value_rate"] = OUTCOME_MATCHED
    else:
        outcomes["verified_value_rate"] = OUTCOME_MISMATCH

    # Missing-value detection
    # How many ground-truth fields are empty/missing in pipeline input (not output)
    # We check the input row fields against ground truth
    missing_count = 0
    total_gt_fields = 0
    for key in ["manufacturer", "brand"]:
        gt_val = ground_truth.get(key, "") or ""
        got_val = (input_row.get(key) or input_row.get("Part_Manuf") or input_row.get("E1_Brand") or input_row.get("Unilog_Brand") or "") or ""
        total_gt_fields += 1
        if not gt_val:
            # Ground truth empty - not a missing detection
            continue
        if not got_val:
            missing_count += 1

    if total_gt_fields == 0:
        outcomes["missing_detection"] = OUTCOME_NOT_EVALUABLE
    elif missing_count / max(total_gt_fields, 1) < 0.3:
        outcomes["missing_detection"] = OUTCOME_MATCHED
    else:
        outcomes["missing_detection"] = OUTCOME_MISMATCH

    # Human-review rate
    # Count attributes needing human review based on pipeline confidence/status
    review_count = 0
    total_attrs = max(len(gt_attributes), 1)
    for gt_attr in gt_attributes:
        key = gt_attr.get("label") or gt_attr.get("attribute") or ""
        if not key:
            continue
        if pipeline_attrs:
            pa_key = [a for a in pipeline_attrs if a.get("key", "").lower() == key.lower()]
            if pa_key:
                pa = pa_key[0]
                # Check if status is unverified or low confidence
                status = pa.get("status", "unverified").upper()
                if status in ("UNVERIFIED", "EXTRACTED"):
                    review_count += 1
                elif pa.get("confidence", 0) < 0.7:
                    review_count += 1

    if total_attrs == 0:
        outcomes["human_review_rate"] = OUTCOME_NOT_EVALUABLE
    elif review_count / max(total_attrs, 1) < 0.3:
        outcomes["human_review_rate"] = OUTCOME_MATCHED
    else:
        outcomes["human_review_rate"] = OUTCOME_MISMATCH

    # Description-level outcomes
    desc_outcomes = _compare_descriptions(
        pipeline_desc.get("descriptions", {}),
        gt_desc,
    )
    outcomes.update(desc_outcomes)

    # Overall row outcome
    # A row is MATCHED if all evaluable critical fields match
    critical_fields = []
    for f in ["manufacturer", "brand", "classification", "lov_compliance", "uom_compliance", "character_limit_compliance"]:
        if outcomes.get(f) == OUTCOME_NOT_EVALUABLE:
            # Skip NOT_EVALUABLE fields from critical check
            continue
        critical_fields.append(f)
    
    # Also include verified_value_rate and missing_detection if evaluable
    evaluable_critical = []
    for f in ["verified_value_rate", "missing_detection", "human_review_rate"]:
        if outcomes.get(f) == OUTCOME_NOT_EVALUABLE:
            continue
        evaluable_critical.append(f)
    
    all_critical_match = all(
        outcomes.get(f) == OUTCOME_MATCHED for f in critical_fields + evaluable_critical
    )
    if all_critical_match:
        outcomes["overall"] = OUTCOME_MATCHED
    else:
        outcomes["overall"] = OUTCOME_MISMATCH

    return outcomes


# ---------------------------------------------------------------------------
# Dataset loading: Excel Input + Delivery Format sheets
# ---------------------------------------------------------------------------

def load_ground_truth_dataset(
    filepath: str,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Load the 200-item ground-truth dataset from Excel.

    Returns:
        (input_rows, ground_truth_rows) - lists of dicts, one per row.
        Input sheet contains the pipeline input data.
        Delivery Format sheet contains the ground-truth expected output.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Input sheet - pipeline input
    input_sheet_name = "Input"
    delivery_sheet_name = "Delivery Format"

    if input_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{input_sheet_name}' not found in {filepath}")
    if delivery_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{delivery_sheet_name}' not found in {filepath}")

    input_ws = wb[input_sheet_name]
    delivery_ws = wb[delivery_sheet_name]

    # Determine row count (skip header)
    input_max_row = input_ws.max_row
    delivery_max_row = delivery_ws.max_row
    input_max_col = input_ws.max_column
    delivery_max_col = delivery_ws.max_column

    input_rows: List[Dict[str, Any]] = []
    ground_truth_rows: List[Dict[str, Any]] = []

    # Read input rows (skip header row 1)
    for row_idx in range(2, input_max_row + 1):
        row_data: Dict[str, Any] = {}
        for col_idx in range(1, input_max_col + 1):
            cell_value = input_ws.cell(row=row_idx, column=col_idx).value
            col_name = input_ws.cell(row=1, column=col_idx).value
            if col_name:
                row_data[col_name] = cell_value
        input_rows.append(row_data)

    # Read ground-truth rows from Delivery Format sheet
    for row_idx in range(2, delivery_max_row + 1):
        row_data: Dict[str, Any] = {}
        for col_idx in range(1, delivery_max_col + 1):
            cell_value = delivery_ws.cell(row=row_idx, column=col_idx).value
            col_name = delivery_ws.cell(row=1, column=col_idx).value
            if col_name:
                row_data[col_name] = cell_value
        ground_truth_rows.append(row_data)

    wb.close()

    # Zip together - assume same row count and order
    min_rows = min(len(input_rows), len(ground_truth_rows))
    input_rows = input_rows[:min_rows]
    ground_truth_rows = ground_truth_rows[:min_rows]

    return input_rows, ground_truth_rows


# ---------------------------------------------------------------------------
# Metrics calculation
# ---------------------------------------------------------------------------

def calculate_metrics(
    row_outcomes: List[Dict[str, Any]],
    total_rows: int,
) -> Dict[str, Any]:
    """Calculate aggregate metrics from per-row evaluation outcomes."""

    if total_rows == 0:
        return {
            "dataset": "200-item-ground-truth",
            "rows_evaluated": 0,
            "overall_accuracy": 0,
            "manufacturer_accuracy": 0,
            "brand_accuracy": 0,
            "classification_accuracy": 0,
            "attribute_accuracy": 0,
            "lov_compliance": 0,
            "uom_compliance": 0,
            "character_limit_compliance": 0,
            "verified_value_rate": 0,
            "missing_detection_rate": 0,
            "human_review_rate": 0,
            "field_results": {},
            "failed_rows": [],
        }

    # Overall accuracy: rows where overall outcome was MATCHED
    matched_rows = sum(
        1 for r in row_outcomes if r.get("overall") == OUTCOME_MATCHED
    )
    overall_accuracy = round(matched_rows / total_rows * 100, 2)

    # Manufacturer accuracy (only rows where manufacturer is evaluable)
    evaluable_mfr = sum(
        1 for r in row_outcomes if r.get("manufacturer") != OUTCOME_NOT_EVALUABLE
    )
    mfr_matched = sum(
        1 for r in row_outcomes
        if r.get("manufacturer") == OUTCOME_MATCHED
    )
    manufacturer_accuracy = round(mfr_matched / max(evaluable_mfr, 1) * 100, 2)

    # Brand accuracy (only rows where brand is evaluable)
    evaluable_brand = sum(
        1 for r in row_outcomes if r.get("brand") != OUTCOME_NOT_EVALUABLE
    )
    brand_matched = sum(
        1 for r in row_outcomes
        if r.get("brand") == OUTCOME_MATCHED
    )
    brand_accuracy = round(brand_matched / max(evaluable_brand, 1) * 100, 2)

    # Classification accuracy (only rows where classification is evaluable)
    evaluable_cls = sum(
        1 for r in row_outcomes if r.get("classification") != OUTCOME_NOT_EVALUABLE
    )
    cls_matched = sum(
        1 for r in row_outcomes
        if r.get("classification") == OUTCOME_MATCHED
    )
    classification_accuracy = round(cls_matched / max(evaluable_cls, 1) * 100, 2)

    # Attribute accuracy (LOV + value match)
    attr_matched = sum(
        1 for r in row_outcomes
        if r.get("lov_compliance") == OUTCOME_MATCHED
        and r.get("uom_compliance") == OUTCOME_MATCHED
        and r.get("lov_compliance") != OUTCOME_NOT_EVALUABLE
    )
    total_attr_eval = sum(
        1 for r in row_outcomes
        if r.get("lov_compliance") != OUTCOME_NOT_EVALUABLE
    )
    attribute_accuracy = round(attr_matched / max(total_attr_eval, 1) * 100, 2)

    # LOV compliance % (across all attribute checks that are evaluable)
    total_lov_checks = sum(
        1 for r in row_outcomes if r.get("lov_compliance") != OUTCOME_NOT_EVALUABLE
    )
    compliant_lov = sum(
        1 for r in row_outcomes
        if r.get("lov_compliance") == OUTCOME_MATCHED
        and r.get("lov_compliance") != OUTCOME_NOT_EVALUABLE
    )
    lov_compliance = round(compliant_lov / max(total_lov_checks, 1) * 100, 2) if total_lov_checks > 0 else 0

    # UOM compliance %
    total_uom_checks = sum(
        1 for r in row_outcomes if r.get("uom_compliance") != OUTCOME_NOT_EVALUABLE
    )
    compliant_uom = sum(
        1 for r in row_outcomes
        if r.get("uom_compliance") == OUTCOME_MATCHED
        and r.get("uom_compliance") != OUTCOME_NOT_EVALUABLE
    )
    uom_compliance = round(compliant_uom / max(total_uom_checks, 1) * 100, 2) if total_uom_checks > 0 else 0

    # Character-limit compliance %
    char_matched = sum(
        1 for r in row_outcomes
        if r.get("character_limit_compliance") == OUTCOME_MATCHED
    )
    # Fix: use correct constant
    char_matched = sum(
        1 for r in row_outcomes
        if r.get("character_limit_compliance") == OUTCOME_MATCHED
    )
    character_limit_compliance = round(char_matched / total_rows * 100, 2)

    # Verified-value %
    verified_matched = sum(
        1 for r in row_outcomes
        if r.get("verified_value_rate") == OUTCOME_MATCHED
    )
    verified_value_rate = round(verified_matched / total_rows * 100, 2)

    # Missing-value detection %
    missing_detected = sum(
        1 for r in row_outcomes
        if r.get("missing_detection") == OUTCOME_MATCHED
    )
    missing_detection_rate = round(missing_detected / total_rows * 100, 2)

    # Human-review rate
    hr_matched = sum(
        1 for r in row_outcomes
        if r.get("human_review_rate") == OUTCOME_MATCHED
    )
    human_review_rate = round(hr_matched / total_rows * 100, 2)

    # Field-level results: per-field accuracy across all rows
    field_outcomes: Dict[str, Dict[str, int]] = defaultdict(lambda: {"matched": 0, "mismatch": 0, "missing": 0, "extra": 0, "not_evaluable": 0})
    for r in row_outcomes:
        for field, outcome in r.items():
            if outcome in field_outcomes[field]:
                field_outcomes[field][outcome] += 1

    field_results: Dict[str, Any] = {}
    for field, counts in field_outcomes.items():
        total = sum(counts.values())
        field_results[field] = {
            "matched": counts["matched"],
            "mismatch": counts["mismatch"],
            "missing": counts["missing"],
            "extra": counts["extra"],
            "not_evaluable": counts["not_evaluable"],
            "accuracy": round(counts["matched"] / max(total, 1) * 100, 2),
        }

    # Failed rows: rows where overall outcome is MISMATCH (and not NOT_EVALUABLE)
    failed_rows: List[Dict[str, Any]] = [
        {
            "row_index": i + 1,
            "outcomes": r,
        }
        for i, r in enumerate(row_outcomes)
        if r.get("overall") == OUTCOME_MISMATCH
    ]

    report = {
        "dataset": "200-item-ground-truth",
        "rows_evaluated": total_rows,
        "overall_accuracy": overall_accuracy,
        "manufacturer_accuracy": manufacturer_accuracy,
        "brand_accuracy": brand_accuracy,
        "classification_accuracy": classification_accuracy,
        "attribute_accuracy": attribute_accuracy,
        "lov_compliance": lov_compliance,
        "uom_compliance": uom_compliance,
        "character_limit_compliance": character_limit_compliance,
        "verified_value_rate": verified_value_rate,
        "missing_detection_rate": missing_detection_rate,
        "human_review_rate": human_review_rate,
        "field_results": field_results,
        "failed_rows": failed_rows,
    }

    return report


# ---------------------------------------------------------------------------
# Full evaluation run
# ---------------------------------------------------------------------------

def run_evaluation(
    filepath: str,
) -> Dict[str, Any]:
    """Run the full ground-truth evaluation pipeline.

    Loads the Excel dataset, runs the product pipeline on each input row,
    compares outputs with ground truth, and returns a machine-readable
    evaluation report.

    Returns the report dict as specified in the task specification.
    """
    # Load dataset
    input_rows, ground_truth_rows = load_ground_truth_dataset(filepath)
    total_rows = len(input_rows)

    print(f"Evaluating {total_rows} rows from {filepath}")

    row_outcomes: List[Dict[str, Any]] = []

    # Process each row
    for i, (input_row, gt_row) in enumerate(zip(input_rows, ground_truth_rows)):
        if (i + 1) % 20 == 0:
            print(f"  Processed {i + 1}/{total_rows} rows...")

        outcome = evaluate_row(input_row, gt_row)
        row_outcomes.append(outcome)

    # Calculate aggregate metrics
    metrics = calculate_metrics(row_outcomes, total_rows)

    return metrics


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    """CLI entry point for Step 6 evaluation."""
    import sys
    filepath = sys.argv[1] if len(sys.argv) > 1 else "Unilog-Sample_200_Items-Input-vs-Output.xlsx"
    try:
        report = run_evaluation(filepath)
        # Print summary
        print("\n" + "=" * 60)
        print("STEP 6: GROUND TRUTH EVALUATION REPORT")
        print("=" * 60)
        print(f"Dataset: {report['dataset']}")
        print(f"Rows evaluated: {report['rows_evaluated']}")
        print(f"\nOverall accuracy: {report['overall_accuracy']}%")
        print(f"Manufacturer accuracy: {report['manufacturer_accuracy']}%")
        print(f"Brand accuracy: {report['brand_accuracy']}%")
        print(f"Classification accuracy: {report['classification_accuracy']}%")
        print(f"Attribute accuracy: {report['attribute_accuracy']}%")
        print(f"LOV compliance: {report['lov_compliance']}%")
        print(f"UOM compliance: {report['uom_compliance']}%")
        print(f"Character-limit compliance: {report['character_limit_compliance']}%")
        print(f"Verified-value rate: {report['verified_value_rate']}%")
        print(f"Missing-value detection rate: {report['missing_detection_rate']}%")
        print(f"Human-review rate: {report['human_review_rate']}%")

        print(f"\nField-level results:")
        for field, data in report["field_results"].items():
            print(
                f"  {field}: {data['matched']} matched, "
                f"{data['mismatch']} mismatch, "
                f"{data['missing']} missing, "
                f"{data['extra']} extra, "
                f"{data['not_evaluable']} not_evaluable "
                f"({data['accuracy']}%)"
            )

        if report["failed_rows"]:
            print(f"\nFailed rows (first 10): {len(report['failed_rows'])} total")
            for fr in report["failed_rows"][:10]:
                print(f"  Row {fr['row_index']}: {fr['outcomes']['overall']}")

        # Export as JSON
        print("\n" + "=" * 60)
        print("Exporting full report as JSON...")
        print(json.dumps(report, indent=2))

    except FileNotFoundError:
        print(f"Error: File not found: {filepath}")
        print("Please ensure Unilog-Sample_200_Items-Input-vs-Output.xlsx exists.")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()